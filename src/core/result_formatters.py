# src/core/result_formatters.py
"""
Formatters for query results
Converts query results to various formats (CSV, Excel, HTML, PDF)
"""

import io
import csv
from typing import List, Dict, Any
from datetime import datetime
import logging

logger = logging.getLogger(__name__)


class QueryResult:
    """Query result container"""

    def __init__(self, columns: List[str], data: List[Dict]):
        self.columns = columns
        self.data = data

    def __len__(self):
        return len(self.data)


class CSVFormatter:
    """Format query results as CSV"""

    def format(self, result: QueryResult) -> bytes:
        """
        Convert query result to CSV bytes

        Args:
            result: QueryResult object

        Returns:
            CSV data as bytes
        """
        try:
            output = io.StringIO()
            writer = csv.writer(output, quoting=csv.QUOTE_MINIMAL)

            # Write headers
            writer.writerow(result.columns)

            # Write data
            for row in result.data:
                # Row can be either list or dict
                if isinstance(row, dict):
                    # Dict format: use .get()
                    writer.writerow([
                        self._format_value(row.get(col, ''))
                        for col in result.columns
                    ])
                else:
                    # List format: use by index
                    writer.writerow([
                        self._format_value(val)
                        for val in row
                    ])

            csv_string = output.getvalue()
            logger.info(f"✅ Formatted {len(result.data)} rows to CSV")

            return csv_string.encode('utf-8')

        except Exception as e:
            logger.error(f"❌ CSV formatting error: {e}")
            raise

    def _format_value(self, value: Any) -> str:
        """Format individual value for CSV"""
        if value is None:
            return ''
        if isinstance(value, (list, dict)):
            return str(value)
        return str(value)

    def get_filename(self, schedule_name: str) -> str:
        """Generate filename"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        safe_name = "".join(c for c in schedule_name if c.isalnum() or c in (' ', '_')).strip()
        safe_name = safe_name.replace(' ', '_')
        return f"{safe_name}_{timestamp}.csv"


class ExcelFormatter:
    """Format query results as Excel"""

    def format(self, result: QueryResult) -> bytes:
        """
        Convert query result to Excel bytes

        Args:
            result: QueryResult object

        Returns:
            Excel data as bytes
        """
        try:
            # Check if openpyxl is available
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            except ImportError:
                logger.warning("⚠️  openpyxl not installed, falling back to CSV")
                return CSVFormatter().format(result)

            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Query Results"

            # Header styling
            header_fill = PatternFill(
                start_color="D4A574",
                end_color="D4A574",
                fill_type="solid"
            )
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_alignment = Alignment(horizontal="left", vertical="center")

            border_side = Side(style='thin', color='CCCCCC')
            border = Border(
                left=border_side,
                right=border_side,
                top=border_side,
                bottom=border_side
            )

            # Write headers
            for col_idx, column in enumerate(result.columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=column)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border

            # Write data
            for row_idx, row in enumerate(result.data, 2):
                for col_idx, column in enumerate(result.columns, 1):
                    # Row can be either list or dict
                    if isinstance(row, dict):
                        value = row.get(column, '')
                    else:
                        # List format: use by index
                        value = row[col_idx - 1] if col_idx - 1 < len(row) else ''

                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = border
                    cell.alignment = Alignment(horizontal="left", vertical="top")

            # Auto-size columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter

                for cell in column:
                    try:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass

                # Set width (min 10, max 50)
                adjusted_width = min(max(max_length + 2, 10), 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Freeze header row
            ws.freeze_panes = ws['A2']

            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            excel_bytes = output.getvalue()

            logger.info(f"✅ Formatted {len(result.data)} rows to Excel ({len(excel_bytes)} bytes)")

            return excel_bytes

        except Exception as e:
            logger.error(f"❌ Excel formatting error: {e}")
            raise

    def get_filename(self, schedule_name: str) -> str:
        """Generate filename"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        safe_name = "".join(c for c in schedule_name if c.isalnum() or c in (' ', '_')).strip()
        safe_name = safe_name.replace(' ', '_')
        return f"{safe_name}_{timestamp}.xlsx"


class HTMLFormatter:
    """Format query results as HTML table"""

    def format(self, result: QueryResult) -> str:
        """
        Convert query result to HTML table

        Args:
            result: QueryResult object

        Returns:
            HTML string
        """
        try:
            html = """
            <table style="border-collapse: collapse; width: 100%; font-family: Arial, sans-serif; font-size: 14px;">
                <thead>
                    <tr style="background: linear-gradient(135deg, #D4A574 0%, #A67C52 100%);">
            """

            # Headers
            for column in result.columns:
                html += f'''
                    <th style="padding: 12px; color: white; text-align: left; 
                               border: 1px solid #999; font-weight: 600;">
                        {self._escape_html(str(column))}
                    </th>
                '''

            html += "</tr></thead><tbody>"

            # Data rows
            for idx, row in enumerate(result.data):
                bg_color = "#f9f9f9" if idx % 2 == 0 else "#ffffff"
                html += f'<tr style="background: {bg_color};">'

                for col_idx, column in enumerate(result.columns):
                    # Row can be either list or dict
                    if isinstance(row, dict):
                        value = row.get(column, '')
                    else:
                        # List format: use by index
                        value = row[col_idx] if col_idx < len(row) else ''

                    formatted_value = self._format_value(value)

                    html += f'''
                        <td style="padding: 10px; border: 1px solid #ddd; 
                                   color: #333; vertical-align: top;">
                            {self._escape_html(formatted_value)}
                        </td>
                    '''

                html += "</tr>"

            html += "</tbody></table>"

            logger.info(f"✅ Formatted {len(result.data)} rows to HTML")

            return html

        except Exception as e:
            logger.error(f"❌ HTML formatting error: {e}")
            raise

    def _escape_html(self, text: str) -> str:
        """Escape HTML special characters"""
        return (str(text)
                .replace('&', '&amp;')
                .replace('<', '&lt;')
                .replace('>', '&gt;')
                .replace('"', '&quot;')
                .replace("'", '&#39;'))

    def _format_value(self, value: Any) -> str:
        """Format individual value"""
        if value is None:
            return ''
        if isinstance(value, (list, dict)):
            return str(value)
        return str(value)

    def get_filename(self, schedule_name: str) -> str:
        """Generate filename"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        safe_name = "".join(c for c in schedule_name if c.isalnum() or c in (' ', '_')).strip()
        safe_name = safe_name.replace(' ', '_')
        return f"{safe_name}_{timestamp}.html"


class ResultFormatterFactory:
    """Factory for creating formatters"""

    _formatters = {
        'csv': CSVFormatter,
        'excel': ExcelFormatter,
        'html': HTMLFormatter
    }

    @classmethod
    def get_formatter(cls, format_type: str):
        """
        Get formatter by type

        Args:
            format_type: 'csv', 'excel', or 'html'

        Returns:
            Formatter instance
        """
        formatter_class = cls._formatters.get(format_type.lower())

        if not formatter_class:
            logger.warning(f"⚠️  Unknown format '{format_type}', using CSV")
            formatter_class = CSVFormatter

        return formatter_class()

    @classmethod
    def get_supported_formats(cls) -> List[str]:
        """Get list of supported formats"""
        return list(cls._formatters.keys())


# Helper function for easy use
def format_result(
        columns: List[str],
        data: List[Dict],
        format_type: str = 'excel',
        schedule_name: str = 'query_result'
) -> Dict[str, Any]:
    """
    Format query result to specified format

    Args:
        columns: Column names
        data: Row data as list of dicts
        format_type: Output format ('csv', 'excel', 'html')
        schedule_name: Name for filename generation

    Returns:
        Dict with 'content' (bytes/str) and 'filename'
    """
    result = QueryResult(columns, data)
    formatter = ResultFormatterFactory.get_formatter(format_type)

    content = formatter.format(result)
    filename = formatter.get_filename(schedule_name)

    return {
        'content': content,
        'filename': filename,
        'format': format_type
    }